"""
pages/7_Valuation_Tracker.py
Price vs EPS Growth Comparison across 25 indices.

Display modes:
  1. Excel-replica yearly + quarterly tables (matches original .xlsx exactly)
  2. Price CAGR vs EPS CAGR divergence chart
  3. Heatmap across all 25 indices
  4. Excel export button (rebuilds the .xlsx)
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import os, sys, json, io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from theme import (
    apply_theme, render_header,
    RH_MAROON, RH_MAROON_DK, RH_GOLD, RH_GOLD_LIGHT, RH_GOLD_DIM,
    RH_RED, RH_GREEN, RH_BG, RH_SURFACE, RH_SURFACE2, RH_TEXT, RH_MUTED, RH_BORDER
)
from index_config import INDICES, BY_SHEET

st.set_page_config(layout="wide", page_title="RH | Valuation Tracker",
                   initial_sidebar_state="expanded")
apply_theme()
render_header("Scanner 07 · Price vs EPS Growth")


# ─────────────────────────────────────────────────────────────────────────────
# Data load
# ─────────────────────────────────────────────────────────────────────────────

JSON_CANDIDATES = [
    os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "valuation_history.json"),
    os.path.join(os.getcwd(), "data", "valuation_history.json"),
    os.path.join(os.getcwd(), "valuation_history.json"),
]

@st.cache_data(ttl=1800, show_spinner=False)
def load_data():
    for p in JSON_CANDIDATES:
        if os.path.exists(p):
            with open(p) as f:
                return json.load(f), p
    return None, None


# ─────────────────────────────────────────────────────────────────────────────
# Compute helpers
# ─────────────────────────────────────────────────────────────────────────────

def yearly_snapshots(snapshots, target_years):
    """Pick the snapshot closest to each year-end (and the latest one)."""
    if not snapshots:
        return []
    snaps = sorted(snapshots, key=lambda s: s["date"])
    out = []
    for y in target_years:
        # find last snap in year y
        year_snaps = [s for s in snaps if s["date"].startswith(str(y))]
        if year_snaps:
            out.append((f"{y}-12-31", year_snaps[-1]))
    # Latest
    latest = snaps[-1]
    if not out or out[-1][1]["date"] != latest["date"]:
        out.append((latest["date"], latest))
    return out


def quarterly_snapshots(snapshots, start_year=2023):
    """One snapshot per quarter end + latest."""
    if not snapshots:
        return []
    snaps = sorted(snapshots, key=lambda s: s["date"])
    out = []
    seen = set()
    for s in snaps:
        y = int(s["date"][:4])
        m = int(s["date"][5:7])
        if y < start_year:
            continue
        q = (m - 1) // 3 + 1
        key = (y, q)
        if key not in seen:
            seen.add(key)
            out.append(s)
        else:
            # replace with later-in-quarter snap
            out[-1] = s
    # Always include latest
    if out and out[-1]["date"] != snaps[-1]["date"]:
        out.append(snaps[-1])
    return out


def cagr(end, start, years):
    if not start or start <= 0 or years <= 0:
        return None
    return (end / start) ** (1 / years) - 1


def compute_yearly_growth(yearly):
    """Returns list of dicts with: label, close, pe, pb, eps, rsi, years_from_base, yoy_idx, yoy_eps, cum_idx, cum_eps."""
    if not yearly:
        return []
    base_label, base_snap = yearly[0]
    base_dt = pd.Timestamp(base_snap["date"])
    rows = []
    for i, (label, snap) in enumerate(yearly):
        dt = pd.Timestamp(snap["date"])
        years = (dt - base_dt).days / 365.25 if i > 0 else 0
        row = {
            "label":     label[:7] if "-" in label else label,
            "date":      snap["date"],
            "close":     snap["close"],
            "pe":        snap["pe"],
            "pb":        snap["pb"],
            "eps":       snap["eps"],
            "rsi":       snap["rsi"],
            "years":     round(years, 2),
            "yoy_idx":   None,
            "yoy_eps":   None,
            "cum_idx":   None,
            "cum_eps":   None,
        }
        if i > 0:
            prev = yearly[i-1][1]
            prev_dt = pd.Timestamp(prev["date"])
            yr_step = (dt - prev_dt).days / 365.25
            row["yoy_idx"] = cagr(snap["close"], prev["close"], yr_step) if prev["close"] else None
            row["yoy_eps"] = cagr(snap["eps"], prev["eps"], yr_step) if (snap["eps"] and prev["eps"]) else None
            row["cum_idx"] = cagr(snap["close"], base_snap["close"], years) if base_snap["close"] else None
            row["cum_eps"] = cagr(snap["eps"], base_snap["eps"], years) if (snap["eps"] and base_snap["eps"]) else None
        rows.append(row)
    return rows


def compute_quarterly_growth(quarterly):
    if not quarterly:
        return []
    rows = []
    for i, snap in enumerate(quarterly):
        row = {
            "label":     snap["date"][:7],
            "date":      snap["date"],
            "close":     snap["close"],
            "eps":       snap["eps"],
            "qoq_idx":   None,
            "qoq_eps":   None,
            "yoy_idx":   None,
            "yoy_eps":   None,
        }
        if i > 0:
            prev = quarterly[i-1]
            if prev["close"]:
                row["qoq_idx"] = (snap["close"] - prev["close"]) / prev["close"] * 100
            if prev["eps"] and snap["eps"]:
                row["qoq_eps"] = (snap["eps"] - prev["eps"]) / prev["eps"] * 100
        if i >= 4:
            prev_yr = quarterly[i-4]
            if prev_yr["close"]:
                row["yoy_idx"] = (snap["close"] - prev_yr["close"]) / prev_yr["close"] * 100
            if prev_yr["eps"] and snap["eps"]:
                row["yoy_eps"] = (snap["eps"] - prev_yr["eps"]) / prev_yr["eps"] * 100
        rows.append(row)
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Page CSS
# ─────────────────────────────────────────────────────────────────────────────

st.markdown(f"""
<style>
.kpi-row {{
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 10px;
    margin-bottom: 22px;
}}
.kpi {{
    background: #FFFFFF;
    border: 1px solid {RH_BORDER};
    border-top: 3px solid {RH_MAROON};
    padding: 14px 16px;
    text-align: center;
}}
.kpi-val {{
    font-family: 'Fraunces', serif;
    font-size: 1.9rem;
    font-weight: 900;
    line-height: 1;
}}
.kpi-lbl {{
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.62rem;
    color: {RH_MUTED};
    text-transform: uppercase;
    letter-spacing: 0.16em;
    margin-top: 8px;
}}
.val-table {{
    width: 100%;
    border-collapse: collapse;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 11px;
    background: {RH_SURFACE};
    margin-bottom: 24px;
}}
.val-table th {{
    background: {RH_MAROON};
    color: {RH_BG};
    text-align: right;
    padding: 8px 10px;
    font-weight: 500;
    letter-spacing: 0.04em;
    border-right: 1px solid rgba(255,255,255,0.1);
}}
.val-table th:first-child {{
    text-align: left;
    background: {RH_MAROON_DK};
}}
.val-table td {{
    padding: 7px 10px;
    text-align: right;
    color: {RH_TEXT};
    border-bottom: 1px solid {RH_BORDER};
    border-right: 1px solid {RH_BORDER};
}}
.val-table td:first-child {{
    text-align: left;
    color: {RH_GOLD_DIM};
    font-weight: 500;
    background: {RH_SURFACE2};
}}
.val-table tr:hover td {{ background: rgba(212,168,48,0.08); }}
.val-table .green {{ color: {RH_GREEN}; font-weight: 500; }}
.val-table .red   {{ color: {RH_RED};   font-weight: 500; }}
.val-table .muted {{ color: {RH_MUTED}; }}
.section-title {{
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.72rem;
    color: {RH_GOLD_DIM};
    margin: 26px 0 12px;
    letter-spacing: 0.16em;
    text-transform: uppercase;
}}
.section-title .accent {{ color: {RH_GOLD_LIGHT}; font-weight: 500; }}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# Load data
# ─────────────────────────────────────────────────────────────────────────────

data, data_path = load_data()
if data is None:
    st.error("No valuation history data found. Run `update_valuation.py` and `reconstruct_valuation_history.py` first.")
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────────────────────────────────────

st.sidebar.markdown(f"""<div style='color:{RH_GOLD_LIGHT};font-size:12px;
letter-spacing:0.2em;text-transform:uppercase;margin-bottom:14px;'>
Valuation Tracker</div>""", unsafe_allow_html=True)

view = st.sidebar.radio(
    "VIEW",
    ["Single Index", "Cross-Index Heatmap", "Divergence Chart", "QoQ Growth", "Sector Rotation", "Screener"],
    index=0,
)

available = [ix["sheet_name"] for ix in INDICES if ix["sheet_name"] in data["indices"]]
default_idx = "NIFTY 50" if "NIFTY 50" in available else (available[0] if available else None)

if view == "Single Index":
    selected_sheet = st.sidebar.selectbox("INDEX", available, index=available.index(default_idx) if default_idx else 0)
    history_range = st.sidebar.radio("YEARLY HISTORY", ["Last 7y", "Last 10y", "Full (since inception)"], index=0)
    quarter_range = st.sidebar.radio("QUARTERLY HISTORY", ["Last 12 quarters", "Last 20 quarters", "Full"], index=0)
else:
    history_range = "Last 7y"
    quarter_range = "Last 12 quarters"

st.sidebar.markdown("---")
if st.sidebar.button("⟳ REFRESH DATA"):
    st.cache_data.clear()
    st.rerun()

last_update = data.get("last_updated", "")
if last_update:
    try:
        ts = datetime.fromisoformat(last_update)
        st.sidebar.markdown(f"<div style='font-size:9px;color:{RH_GOLD_LIGHT};margin-top:10px;'>"
                            f"Last update<br>{ts:%d-%b-%Y %H:%M}</div>", unsafe_allow_html=True)
    except: pass


# ─────────────────────────────────────────────────────────────────────────────
# Format helpers
# ─────────────────────────────────────────────────────────────────────────────

def fmt_num(v, dp=2):
    if v is None or pd.isna(v):
        return "<span class='muted'>—</span>"
    return f"{v:,.{dp}f}"


def fmt_pct(v, dp=2):
    if v is None or pd.isna(v):
        return "<span class='muted'>—</span>"
    cls = "green" if v >= 0 else "red"
    return f"<span class='{cls}'>{v*100:+.{dp}f}%</span>"


def fmt_pct_raw(v, dp=2):
    """Already a percent number (e.g. 5.2 means 5.2%)."""
    if v is None or pd.isna(v):
        return "<span class='muted'>—</span>"
    cls = "green" if v >= 0 else "red"
    return f"<span class='{cls}'>{v:+.{dp}f}%</span>"


# ─────────────────────────────────────────────────────────────────────────────
# VIEW 1 — Single Index
# ─────────────────────────────────────────────────────────────────────────────

def render_single_index(sheet_name):
    ix = BY_SHEET[sheet_name]
    snaps = data["indices"][sheet_name]["snapshots"]
    if not snaps:
        st.warning(f"No snapshots for {sheet_name}")
        return

    # Build yearly + quarterly views — use index's actual start_year for inception view
    start_yr = ix.get("start_year", 2012)
    years = list(range(start_yr, datetime.now().year))
    yearly_full = yearly_snapshots(snaps, years)
    yearly_full_rows = compute_yearly_growth(yearly_full)  # full history — keeps base = start_year

    quarterly_full = quarterly_snapshots(snaps, start_year=start_yr)
    quarterly_full_rows = compute_quarterly_growth(quarterly_full)

    # Apply UI-level windowing (display only — CAGR still computed from full base)
    if history_range == "Last 7y":
        yearly_rows = yearly_full_rows[-8:] if len(yearly_full_rows) > 8 else yearly_full_rows
    elif history_range == "Last 10y":
        yearly_rows = yearly_full_rows[-11:] if len(yearly_full_rows) > 11 else yearly_full_rows
    else:
        yearly_rows = yearly_full_rows

    if quarter_range == "Last 12 quarters":
        quarterly_rows = quarterly_full_rows[-12:]
    elif quarter_range == "Last 20 quarters":
        quarterly_rows = quarterly_full_rows[-20:]
    else:
        quarterly_rows = quarterly_full_rows

    latest = snaps[-1]
    latest_yr = yearly_full_rows[-1] if yearly_full_rows else {}

    # ── KPI row ──
    div = None
    if latest_yr.get("cum_idx") is not None and latest_yr.get("cum_eps") is not None:
        div = latest_yr["cum_idx"] - latest_yr["cum_eps"]
    div_color = RH_GREEN if (div is not None and div <= 0) else RH_RED
    div_str = f"{div*100:+.1f}%" if div is not None else "—"

    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi">
            <div class="kpi-val" style="color:{RH_TEXT}">{latest['close']:,.0f}</div>
            <div class="kpi-lbl">{ix['display']} · Close</div>
        </div>
        <div class="kpi">
            <div class="kpi-val" style="color:{RH_GOLD_LIGHT}">{latest['pe'] if latest['pe'] else '—'}</div>
            <div class="kpi-lbl">P/E Ratio</div>
        </div>
        <div class="kpi">
            <div class="kpi-val" style="color:{RH_MAROON}">{latest['eps'] if latest['eps'] else '—'}</div>
            <div class="kpi-lbl">EPS (derived)</div>
        </div>
        <div class="kpi">
            <div class="kpi-val" style="color:{RH_TEXT}">{latest['rsi'] if latest['rsi'] else '—'}</div>
            <div class="kpi-lbl">RSI (14)</div>
        </div>
        <div class="kpi">
            <div class="kpi-val" style="color:{div_color}">{div_str}</div>
            <div class="kpi-lbl">Price − EPS CAGR Spread</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Yearly performance table (matches Excel exactly) ──
    st.markdown(f"""<div class='section-title'>
        <span class='accent'>Table</span> &nbsp; {ix['display']} — Yearly Performance &amp; CAGR
    </div>""", unsafe_allow_html=True)

    if yearly_rows:
        cols = [r["label"] for r in yearly_rows]
        rows_def = [
            (f"{ix['display']}", "close", fmt_num, 0),
            ("PB",  "pb",  fmt_num, 2),
            ("EPS", "eps", fmt_num, 2),
            ("PE",  "pe",  fmt_num, 2),
            ("RSI", "rsi", fmt_num, 2),
        ]
        html = "<table class='val-table'><thead><tr><th>Metric</th>"
        for c in cols:
            html += f"<th>{c}</th>"
        html += "</tr></thead><tbody>"
        for label, key, fn, dp in rows_def:
            html += f"<tr><td>{label}</td>"
            for r in yearly_rows:
                html += f"<td>{fn(r[key], dp)}</td>"
            html += "</tr>"
        # Spacer + CAGR rows
        html += "<tr><td colspan='99' style='background:transparent;border:none;padding:6px;'></td></tr>"
        html += "<tr><td>No. of Years</td>"
        for r in yearly_rows:
            html += f"<td>{fmt_num(r['years'], 2) if r['years'] else '<span class=muted>—</span>'}</td>"
        html += "</tr>"

        growth_rows = [
            (f"{ix['display']} YOY CAGR", "yoy_idx"),
            ("EPS YOY CAGR",              "yoy_eps"),
            (f"{ix['display']} CUM CAGR", "cum_idx"),
            ("EPS CUM CAGR",              "cum_eps"),
        ]
        for label, key in growth_rows:
            html += f"<tr><td>{label}</td>"
            for r in yearly_rows:
                html += f"<td>{fmt_pct(r[key], 2)}</td>"
            html += "</tr>"
        html += "</tbody></table>"
        st.markdown(html, unsafe_allow_html=True)

    # ── Yearly chart: Price CAGR vs EPS CAGR ──
    st.markdown(f"""<div class='section-title'>
        <span class='accent'>Chart</span> &nbsp; Cumulative CAGR — Price vs EPS
    </div>""", unsafe_allow_html=True)

    chart_rows = [r for r in yearly_rows if r["cum_idx"] is not None or r["cum_eps"] is not None]
    if chart_rows:
        x = [r["label"] for r in chart_rows]
        y_price = [(r["cum_idx"] * 100) if r["cum_idx"] is not None else None for r in chart_rows]
        y_eps   = [(r["cum_eps"] * 100) if r["cum_eps"] is not None else None for r in chart_rows]
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=x, y=y_price, mode="lines+markers",
                                  line=dict(color=RH_MAROON, width=3),
                                  marker=dict(size=8, color=RH_MAROON),
                                  name=f"{ix['display']} CAGR"))
        fig.add_trace(go.Scatter(x=x, y=y_eps, mode="lines+markers",
                                  line=dict(color=RH_GOLD, width=3, dash="dot"),
                                  marker=dict(size=8, color=RH_GOLD),
                                  name="EPS CAGR"))
        fig.update_layout(
            plot_bgcolor=RH_BG, paper_bgcolor=RH_SURFACE2,
            font=dict(family="IBM Plex Mono", color="#2C1810", size=11),
            hoverlabel=dict(bgcolor="#FFFFFF", bordercolor=RH_MAROON,
                            font=dict(family="IBM Plex Mono", color=RH_TEXT)),
            height=380, margin=dict(l=40, r=20, t=20, b=40),
            yaxis=dict(title="CAGR (%)", tickformat=".1f",
                       tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810"),
                       title_font=dict(family="IBM Plex Mono", size=11, color="#2C1810")),
            xaxis=dict(tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810")),
            legend=dict(
                orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                font=dict(family="IBM Plex Mono", size=10, color="#2C1810"),
            ),
        )
        fig.update_xaxes(gridcolor='rgba(139,26,26,0.1)', zeroline=False)
        fig.update_yaxes(gridcolor='rgba(139,26,26,0.1)', zeroline=True, zerolinecolor="#8B6A4A")
        st.plotly_chart(fig, use_container_width=True)

    # ── P/E ratio band chart (mean ± 1 SD) ──
    pe_snaps = [s for s in snaps if s.get("pe") is not None and s["pe"] > 0]
    if len(pe_snaps) >= 4:
        st.markdown(f"""<div class='section-title'>
            <span class='accent'>Chart</span> &nbsp; P/E vs Historical Mean (±1 SD bands)
        </div>""", unsafe_allow_html=True)

        pe_df = pd.DataFrame([{"date": pd.Timestamp(s["date"]), "pe": s["pe"]} for s in pe_snaps]).sort_values("date")
        pe_mean = pe_df["pe"].mean()
        pe_std  = pe_df["pe"].std() if len(pe_df) > 1 else 0
        upper_1 = pe_mean + pe_std
        lower_1 = pe_mean - pe_std
        current_pe = pe_df["pe"].iloc[-1]
        z = (current_pe - pe_mean) / pe_std if pe_std > 0 else 0

        if z > 1:
            verdict, vcolor = "EXPENSIVE", RH_RED
        elif z < -1:
            verdict, vcolor = "CHEAP", RH_GREEN
        else:
            verdict, vcolor = "FAIR", RH_GOLD_DIM

        pe_fig = go.Figure()
        # Fill area for ±1 SD band
        pe_fig.add_trace(go.Scatter(
            x=list(pe_df["date"]) + list(pe_df["date"][::-1]),
            y=[upper_1] * len(pe_df) + [lower_1] * len(pe_df),
            fill="toself", fillcolor="rgba(212,168,48,0.15)",
            line=dict(color="rgba(0,0,0,0)"), hoverinfo="skip", showlegend=False,
        ))
        # Mean line
        pe_fig.add_trace(go.Scatter(
            x=pe_df["date"], y=[pe_mean] * len(pe_df),
            mode="lines", line=dict(color=RH_GOLD_DIM, dash="dash", width=1),
            name=f"Mean ({pe_mean:.1f})",
            hovertemplate=f"Mean P/E: {pe_mean:.2f}<extra></extra>",
        ))
        # Actual P/E line
        pe_fig.add_trace(go.Scatter(
            x=pe_df["date"], y=pe_df["pe"],
            mode="lines+markers",
            line=dict(color=RH_MAROON, width=2.5),
            marker=dict(size=6, color=RH_MAROON),
            name="P/E",
            hovertemplate="%{x|%b %Y}<br>P/E: %{y:.2f}<extra></extra>",
        ))
        # Highlight current point
        pe_fig.add_trace(go.Scatter(
            x=[pe_df["date"].iloc[-1]], y=[current_pe],
            mode="markers", marker=dict(size=14, color=vcolor, line=dict(color="white", width=2)),
            hovertemplate=f"Current: {current_pe:.2f}<br>Z-score: {z:+.2f}<extra></extra>",
            showlegend=False,
        ))
        pe_fig.update_layout(
            plot_bgcolor=RH_BG, paper_bgcolor=RH_SURFACE2,
            font=dict(family="IBM Plex Mono", color="#2C1810", size=11),
            hoverlabel=dict(bgcolor="#FFFFFF", bordercolor=RH_MAROON,
                            font=dict(family="IBM Plex Mono", color=RH_TEXT)),
            height=360, margin=dict(l=40, r=20, t=20, b=40),
            yaxis=dict(
                title="P/E Ratio",
                tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810"),
                title_font=dict(family="IBM Plex Mono", size=11, color="#2C1810"),
            ),
            xaxis=dict(tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810")),
            legend=dict(
                orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                font=dict(family="IBM Plex Mono", size=10, color="#2C1810"),
            ),
        )
        pe_fig.update_xaxes(gridcolor='rgba(139,26,26,0.1)', zeroline=False)
        pe_fig.update_yaxes(gridcolor='rgba(139,26,26,0.1)')
        st.plotly_chart(pe_fig, use_container_width=True)

        # Summary line below chart
        st.markdown(f"""<div style='font-size:11px;color:{RH_TEXT};background:{RH_SURFACE};
        border-left:3px solid {vcolor};padding:10px 14px;margin-top:-8px;'>
        Current P/E <b>{current_pe:.2f}</b> vs historical mean <b>{pe_mean:.2f}</b>
        (σ = {pe_std:.2f}). Z-score <b style='color:{vcolor};'>{z:+.2f}</b> →
        <b style='color:{vcolor};'>{verdict}</b>.
        Based on {len(pe_df)} snapshots from {pe_df['date'].min():%b %Y} to {pe_df['date'].max():%b %Y}.
        </div>""", unsafe_allow_html=True)

    # ── Quarterly table ──
    st.markdown(f"""<div class='section-title'>
        <span class='accent'>Table</span> &nbsp; {ix['display']} — Quarterly Performance
    </div>""", unsafe_allow_html=True)

    if quarterly_rows:
        cols = [r["label"] for r in quarterly_rows]
        html = "<table class='val-table'><thead><tr><th>Metric</th>"
        for c in cols:
            html += f"<th>{c}</th>"
        html += "</tr></thead><tbody>"
        data_rows = [
            (f"{ix['display']}",   "close", fmt_num, 0),
            ("EPS",                "eps",   fmt_num, 2),
        ]
        for label, key, fn, dp in data_rows:
            html += f"<tr><td>{label}</td>"
            for r in quarterly_rows:
                html += f"<td>{fn(r[key], dp)}</td>"
            html += "</tr>"

        html += "<tr><td colspan='99' style='background:transparent;border:none;padding:6px;'></td></tr>"
        growth_rows = [
            ("QOQ Gr (Index)", "qoq_idx"),
            ("QOQ Gr (EPS)",   "qoq_eps"),
            ("YOY Gr (Index)", "yoy_idx"),
            ("YOY Gr (EPS)",   "yoy_eps"),
        ]
        for label, key in growth_rows:
            html += f"<tr><td>{label}</td>"
            for r in quarterly_rows:
                html += f"<td>{fmt_pct_raw(r[key], 2)}</td>"
            html += "</tr>"
        html += "</tbody></table>"
        st.markdown(html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# VIEW 2 — Heatmap
# ─────────────────────────────────────────────────────────────────────────────

def render_heatmap():
    st.markdown(f"""<div class='section-title'>
        <span class='accent'>Heatmap</span> &nbsp; Where Prices Have Outrun Earnings (vs 2019 base)
    </div>""", unsafe_allow_html=True)

    st.markdown(f"""<div style='font-size:11px;color:{RH_TEXT};background:{RH_SURFACE};
    border-left:3px solid {RH_GOLD};padding:12px 16px;margin-bottom:18px;line-height:1.6;'>
    <b style='color:{RH_MAROON};'>How to read this:</b> Each cell shows the
    <b>Price CAGR minus EPS CAGR</b> spread, both measured cumulatively from end-of-2019.
    <br><br>
    &nbsp;&nbsp;<b>+15%</b> means price compounded 15 percentage points <i>faster</i> per year than earnings →
    <span style='color:{RH_RED};'><b>re-rated higher</b></span> (valuation expanded, P/E went up).
    <br>
    &nbsp;&nbsp;<b>−15%</b> means earnings compounded 15 percentage points <i>faster</i> than price →
    <span style='color:{RH_GREEN};'><b>de-rated</b></span> (P/E compressed, potentially undervalued).
    <br>
    &nbsp;&nbsp;<b>~0%</b> means price kept pace with earnings → multiples roughly unchanged.
    <br><br>
    The columns are year-end dates, so the last column shows the spread compounded all the way from 2019 to today —
    that's the cleanest signal of which sectors have re-rated most.
    Indices launched after 2019 (Digital, Defense, Manufacturing) show blank cells until their inception year.
    </div>""", unsafe_allow_html=True)

    # Use a common base year (2019) so all indices are comparable in the heatmap.
    # Indices launched after 2019 just have blank cells for early columns.
    COMMON_BASE = 2019
    rows = []
    for ix in INDICES:
        snaps = data["indices"].get(ix["sheet_name"], {}).get("snapshots", [])
        if not snaps:
            continue
        # Restrict to snapshots from common base onward
        base_year = max(COMMON_BASE, ix.get("start_year", 2012))
        years = list(range(base_year, datetime.now().year))
        yearly = yearly_snapshots(snaps, years)
        yr_rows = compute_yearly_growth(yearly)
        if len(yr_rows) < 2:
            continue
        row_data = {"Index": ix["display"]}
        for r in yr_rows[1:]:  # skip base year (no growth yet)
            if r["cum_idx"] is not None and r["cum_eps"] is not None:
                spread = (r["cum_idx"] - r["cum_eps"]) * 100
                row_data[r["label"]] = round(spread, 1)
        # Only include row if it has at least one data point
        if len(row_data) > 1:
            rows.append(row_data)

    if not rows:
        st.warning("No data for heatmap")
        return

    df = pd.DataFrame(rows).set_index("Index")
    # Order columns chronologically
    df = df[sorted(df.columns)]

    z = df.values
    text = [[f"{v:+.1f}%" if pd.notna(v) else "" for v in row] for row in z]

    fig = go.Figure(data=go.Heatmap(
        z=z, x=list(df.columns), y=list(df.index),
        text=text, texttemplate="%{text}",
        textfont={"family": "IBM Plex Mono", "size": 9, "color": "white"},
        colorscale=[
            [0.0,  RH_GREEN],
            [0.45, "#7FB97F"],
            [0.5,  "#FFFFFF"],
            [0.55, "#E78878"],
            [1.0,  RH_RED],
        ],
        zmid=0,
        hovertemplate="<b>%{y}</b><br>Year: %{x}<br>Spread: %{z:+.1f}%<extra></extra>",
        colorbar=dict(title="Spread %", tickfont=dict(color="#2C1810", family="IBM Plex Mono", size=10)),
    ))
    fig.update_layout(
        plot_bgcolor=RH_BG, paper_bgcolor=RH_SURFACE2,
        font=dict(family="IBM Plex Mono", color="#2C1810", size=10),
        height=max(500, 22 * len(rows)),
        margin=dict(l=140, r=40, t=20, b=40),
        xaxis=dict(tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810")),
        yaxis=dict(tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810")),
    )
    st.plotly_chart(fig, use_container_width=True)

    # Latest leaderboard
    st.markdown(f"""<div class='section-title'>
        <span class='accent'>Leaderboard</span> &nbsp; Latest Spread (Most Overvalued → Most Undervalued)
    </div>""", unsafe_allow_html=True)

    latest_col = df.columns[-1]
    leader = df[[latest_col]].dropna().sort_values(latest_col, ascending=False).rename(columns={latest_col: "Spread %"})
    leader = leader.reset_index()
    leader["Bucket"] = leader["Spread %"].apply(lambda v: "Re-rated (price > EPS)" if v > 5 else ("De-rated (EPS > price)" if v < -5 else "Balanced"))

    html = "<table class='val-table'><thead><tr><th>Rank</th><th>Index</th><th>Spread %</th><th>Bucket</th></tr></thead><tbody>"
    for rank, (_, r) in enumerate(leader.iterrows(), 1):
        cls = "red" if r["Spread %"] > 0 else "green"
        html += f"<tr><td>{rank}</td><td>{r['Index']}</td><td><span class='{cls}'>{r['Spread %']:+.1f}%</span></td><td>{r['Bucket']}</td></tr>"
    html += "</tbody></table>"
    st.markdown(html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# VIEW 3 — Divergence Chart (multi-index)
# ─────────────────────────────────────────────────────────────────────────────

def render_divergence():
    # ── CAGR window selector ──
    head_col, toggle_col = st.columns([3, 4])
    with head_col:
        st.markdown(f"""<div class='section-title'>
            <span class='accent'>Chart</span> &nbsp; Price CAGR vs EPS CAGR
        </div>""", unsafe_allow_html=True)
    with toggle_col:
        st.markdown("<div style='margin-top:18px;'></div>", unsafe_allow_html=True)
        cagr_window = st.radio(
            "CAGR Window",
            ["1Y", "3Y", "5Y", "10Y", "Since 2019"],
            index=2, horizontal=True, label_visibility="collapsed",
            key="divergence_cagr_window",
        )

    # Map window → lookback in years (or fixed base year)
    if cagr_window == "Since 2019":
        lookback_years = None
        window_desc = "since end-of-2019"
    else:
        lookback_years = int(cagr_window.replace("Y", ""))
        window_desc = f"over the last {lookback_years} year{'s' if lookback_years > 1 else ''}"

    st.markdown(f"""<div style='font-size:11px;color:{RH_TEXT};background:{RH_SURFACE};
    border-left:3px solid {RH_GOLD};padding:12px 16px;margin-bottom:18px;line-height:1.6;'>
    <b style='color:{RH_MAROON};'>How to read this:</b> Each dot is one index. Showing CAGR <b>{window_desc}</b>.
    The <b>x-axis</b> is earnings growth (EPS CAGR). The <b>y-axis</b> is price growth (Price CAGR).
    The <b>dashed line is parity</b> — price = earnings growth.
    <br><br>
    &nbsp;&nbsp;<b>Above the line</b> = price re-rated faster than earnings (got more expensive).
    <br>
    &nbsp;&nbsp;<b>Below the line</b> = earnings grew faster than price (got cheaper).
    <br><br>
    Indices without enough history for the selected window are excluded automatically.
    </div>""", unsafe_allow_html=True)

    points = []
    for ix in INDICES:
        snaps = data["indices"].get(ix["sheet_name"], {}).get("snapshots", [])
        if len(snaps) < 2:
            continue
        snaps_sorted = sorted(snaps, key=lambda s: s["date"])
        latest = snaps_sorted[-1]
        latest_dt = pd.Timestamp(latest["date"])

        # Find base snapshot
        if lookback_years is None:
            # Since 2019: find the snapshot closest to (but not after) 2019-12-31
            target_dt = pd.Timestamp("2019-12-31")
        else:
            target_dt = latest_dt - pd.DateOffset(years=lookback_years)

        # Pick the snapshot closest to target_dt. Widened tolerance (180 days) so the
        # sparse seed (year-end snapshots) still produces useful 1Y/3Y/5Y/10Y windows.
        # Once daily updater is running, 90-day tolerance would suffice.
        snap_dates = [pd.Timestamp(s["date"]) for s in snaps_sorted]
        deltas = [abs((d - target_dt).days) for d in snap_dates]
        closest_idx = deltas.index(min(deltas))
        if deltas[closest_idx] > 180:
            # Index doesn't have data far enough back for this window
            continue
        base = snaps_sorted[closest_idx]
        base_dt = pd.Timestamp(base["date"])

        years_diff = (latest_dt - base_dt).days / 365.25
        if years_diff < 0.5:
            continue  # too short to compute meaningful CAGR

        if not base["close"] or not latest["close"] or base["close"] <= 0:
            continue
        if base["eps"] is None or latest["eps"] is None or base["eps"] <= 0:
            continue

        price_cagr = ((latest["close"] / base["close"]) ** (1 / years_diff) - 1) * 100
        eps_cagr   = ((latest["eps"]   / base["eps"])   ** (1 / years_diff) - 1) * 100

        points.append({
            "name": ix["display"],
            "price_cagr": price_cagr,
            "eps_cagr":   eps_cagr,
            "spread":     price_cagr - eps_cagr,
            "base_date":  base["date"],
            "latest_date": latest["date"],
        })

    if not points:
        st.warning(f"No indices have enough history for a {cagr_window} CAGR window.")
        return

    df = pd.DataFrame(points)
    colors = [RH_RED if v > 0 else RH_GREEN for v in df["spread"]]

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df["eps_cagr"], y=df["price_cagr"],
        mode="markers+text",
        marker=dict(size=14, color=colors, line=dict(width=1, color=RH_MAROON_DK)),
        text=df["name"], textposition="top center",
        textfont=dict(family="IBM Plex Mono", size=9, color=RH_TEXT),
        customdata=df[["base_date", "latest_date", "spread"]].values,
        hovertemplate="<b>%{text}</b><br>"
                      "EPS CAGR: %{x:.1f}%<br>"
                      "Price CAGR: %{y:.1f}%<br>"
                      "Spread: %{customdata[2]:+.1f}%<br>"
                      "Base: %{customdata[0]}<br>"
                      "Latest: %{customdata[1]}"
                      "<extra></extra>",
        showlegend=False,
    ))
    # Diagonal y = x line (parity)
    lo = min(df["eps_cagr"].min(), df["price_cagr"].min()) - 5
    hi = max(df["eps_cagr"].max(), df["price_cagr"].max()) + 5
    fig.add_trace(go.Scatter(
        x=[lo, hi], y=[lo, hi], mode="lines",
        line=dict(color=RH_GOLD_DIM, dash="dash", width=1),
        hoverinfo="skip", showlegend=False,
    ))
    fig.add_annotation(x=hi-2, y=hi-1, text="Price = EPS  (fair)",
                       showarrow=False, font=dict(color=RH_GOLD_DIM, size=9, family="IBM Plex Mono"))

    fig.update_layout(
        plot_bgcolor=RH_BG, paper_bgcolor=RH_SURFACE2,
        font=dict(family="IBM Plex Mono", color="#2C1810", size=11),
        height=600, margin=dict(l=60, r=20, t=20, b=50),
        xaxis=dict(
            title="EPS CAGR (%) — earnings growth",
            gridcolor='rgba(139,26,26,0.1)', zeroline=True, zerolinecolor="#8B6A4A",
            tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810"),
            title_font=dict(family="IBM Plex Mono", size=11, color="#2C1810"),
        ),
        yaxis=dict(
            title="Price CAGR (%) — index return",
            gridcolor='rgba(139,26,26,0.1)', zeroline=True, zerolinecolor="#8B6A4A",
            tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810"),
            title_font=dict(family="IBM Plex Mono", size=11, color="#2C1810"),
        ),
    )
    st.plotly_chart(fig, use_container_width=True)

    # ── Ranked summary table ──
    st.markdown(f"""<div class='section-title'>
        <span class='accent'>Ranking</span> &nbsp; {cagr_window} CAGR — Sorted by Spread
    </div>""", unsafe_allow_html=True)

    ranked = df.sort_values("spread", ascending=False).reset_index(drop=True)
    html = "<table class='val-table'><thead><tr><th>Rank</th><th>Index</th>" \
           "<th>Price CAGR</th><th>EPS CAGR</th><th>Spread</th><th>Read</th></tr></thead><tbody>"
    for rank, (_, r) in enumerate(ranked.iterrows(), 1):
        spread = r["spread"]
        if spread > 5:
            read = f"<span class='red'>Re-rated higher (price &gt; EPS)</span>"
        elif spread < -5:
            read = f"<span class='green'>De-rated (EPS &gt; price)</span>"
        else:
            read = "Balanced"
        spread_cls = "red" if spread > 0 else "green"
        html += f"<tr><td>{rank}</td><td>{r['name']}</td>" \
                f"<td>{r['price_cagr']:+.1f}%</td>" \
                f"<td>{r['eps_cagr']:+.1f}%</td>" \
                f"<td><span class='{spread_cls}'>{spread:+.1f}%</span></td>" \
                f"<td>{read}</td></tr>"
    html += "</tbody></table>"
    st.markdown(html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# VIEW 4 — QoQ Growth Heatmap (8 quarters, price / EPS / PE toggle)
# ─────────────────────────────────────────────────────────────────────────────

def render_qoq_heatmap():
    # ── Metric toggle ──
    head_col, toggle_col = st.columns([3, 4])
    with head_col:
        st.markdown(f"""<div class='section-title'>
            <span class='accent'>Heatmap</span> &nbsp; Quarter-on-Quarter Growth — All Sectors
        </div>""", unsafe_allow_html=True)
    with toggle_col:
        st.markdown("<div style='margin-top:18px;'></div>", unsafe_allow_html=True)
        metric = st.radio(
            "Metric",
            ["Price QoQ %", "EPS QoQ %", "P/E QoQ %"],
            index=0, horizontal=True, label_visibility="collapsed",
            key="qoq_metric",
        )

    metric_key = {"Price QoQ %": "close", "EPS QoQ %": "eps", "P/E QoQ %": "pe"}[metric]
    metric_label = metric

    # Color scale clamp — keeps normal cells readable even when outliers exist.
    # Real values always shown as text; cells beyond clamp get ⚡ flag.
    CLAMP = {"Price QoQ %": 30, "EPS QoQ %": 80, "P/E QoQ %": 50}[metric]

    st.markdown(f"""<div style='font-size:11px;color:{RH_TEXT};background:{RH_SURFACE};
    border-left:3px solid {RH_GOLD};padding:12px 16px;margin-bottom:18px;line-height:1.6;'>
    <b style='color:{RH_MAROON};'>How to read this:</b>
    Each cell shows <b>{metric_label}</b> vs the previous quarter. 8 quarters, most recent on the right.
    Rows sorted by latest quarter — hottest at the top.
    <br><br>
    &nbsp;&nbsp;<b style='color:{RH_GREEN};'>Green</b> = grew &nbsp;|&nbsp;
    <b style='color:{RH_RED};'>Red</b> = shrank &nbsp;|&nbsp;
    White = flat &nbsp;|&nbsp;
    <b style='color:#7B2FBE;'>⚡ Purple</b> = extreme outlier (beyond ±{CLAMP}% colour scale — real value shown)
    </div>""", unsafe_allow_html=True)

    # ── Build the matrix ──
    ALL_QUARTERS = set()
    index_data = {}

    for ix in INDICES:
        snaps = data["indices"].get(ix["sheet_name"], {}).get("snapshots", [])
        if not snaps:
            continue
        quarterly = quarterly_snapshots(snaps, start_year=2020)
        if len(quarterly) < 2:
            continue
        row = {}
        for i in range(1, len(quarterly)):
            prev, curr = quarterly[i-1], quarterly[i]
            prev_val = prev.get(metric_key)
            curr_val = curr.get(metric_key)
            if prev_val and curr_val and prev_val > 0 and curr_val is not None:
                pct_chg = (curr_val - prev_val) / abs(prev_val) * 100
                dt = pd.Timestamp(curr["date"])
                q = (dt.month - 1) // 3 + 1
                label = f"{dt.year}-Q{q}"
                row[label] = round(pct_chg, 1)
                ALL_QUARTERS.add(label)
        if row:
            index_data[ix["display"]] = row

    if not index_data:
        st.warning("Not enough quarterly data yet. Run `reconstruct_valuation_history.py` to backfill.")
        return

    all_q_sorted = sorted(ALL_QUARTERS)
    if len(all_q_sorted) > 8:
        all_q_sorted = all_q_sorted[-8:]

    rows = []
    for display_name, row in index_data.items():
        r = {"Index": display_name}
        for q in all_q_sorted:
            r[q] = row.get(q, None)
        non_null = sum(1 for q in all_q_sorted if r.get(q) is not None)
        if non_null >= max(1, len(all_q_sorted) // 2):
            rows.append(r)

    if not rows:
        st.warning("No data.")
        return

    df = pd.DataFrame(rows).set_index("Index")
    df = df[all_q_sorted]

    latest_q = all_q_sorted[-1]
    df = df.sort_values(latest_q, ascending=False, na_position="last")

    # ── Build clamped z matrix + annotated text ──
    # z_clamped: values clipped to ±CLAMP for color only
    # z_real: original values kept for hover
    # text_labels: show real value, flag outliers with ⚡
    z_real    = df.values.tolist()
    z_clamped = []
    text_labels = []

    for row in z_real:
        z_row, t_row = [], []
        for v in row:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                z_row.append(None)
                t_row.append("")
            else:
                is_outlier = abs(v) > CLAMP
                z_row.append(float(np.clip(v, -CLAMP, CLAMP)))
                label = f"⚡{v:+.0f}%" if is_outlier else f"{v:+.1f}%"
                t_row.append(label)
        z_clamped.append(z_row)
        text_labels.append(t_row)

    row_height = 30
    chart_height = max(580, row_height * len(df) + 120)

    fig = go.Figure(data=go.Heatmap(
        z=z_clamped,
        customdata=z_real,
        x=list(df.columns),
        y=list(df.index),
        text=text_labels,
        texttemplate="%{text}",
        textfont={"family": "IBM Plex Mono", "size": 10, "color": "#1A1A1A"},
        colorscale=[
            [0.0,   "#C0392B"],   # strong red
            [0.35,  "#E8A090"],   # light red
            [0.5,   "#F8F4EE"],   # near-white (neutral)
            [0.65,  "#90C090"],   # light green
            [1.0,   "#2E7D32"],   # strong green
        ],
        zmin=-CLAMP,
        zmax=CLAMP,
        hovertemplate=(
            "<b>%{y}</b><br>"
            "%{x}<br>"
            + metric_label + ": <b>%{customdata:+.1f}%</b><br>"
            "<i>(colour clamped at ±" + str(CLAMP) + "%)</i>"
            "<extra></extra>"
        ),
        colorbar=dict(
            title=dict(
                text=f"{metric_label}<br><span style='font-size:8px'>clamped ±{CLAMP}%</span>",
                font=dict(family="IBM Plex Mono", size=10, color="#2C1810"),
            ),
            tickfont=dict(color="#2C1810", family="IBM Plex Mono", size=10),
            tickformat="+d",
            ticksuffix="%",
            thickness=14,
            len=0.85,
            outlinecolor="#2C1810",
            outlinewidth=1,
        ),
    ))
    fig.update_layout(
        plot_bgcolor="#F5ECD7",
        paper_bgcolor="#EDE2C8",
        font=dict(family="IBM Plex Mono", color="#2C1810", size=10),
        height=chart_height,
        margin=dict(l=170, r=80, t=50, b=30),
        xaxis=dict(
            side="top",
            tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810"),
            tickangle=0,
            title_font=dict(color="#2C1810"),
            linecolor="#2C1810",
            gridcolor="rgba(44,24,16,0.1)",
        ),
        yaxis=dict(
            tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810"),
            autorange="reversed",
            title_font=dict(color="#2C1810"),
            linecolor="#2C1810",
            gridcolor="rgba(44,24,16,0.1)",
        ),
    )
    st.plotly_chart(fig, use_container_width=True)

    # ── Summary: best and worst this quarter ──
    st.markdown(f"""<div class='section-title'>
        <span class='accent'>Summary</span> &nbsp; {latest_q} — Best vs Worst ({metric_label})
    </div>""", unsafe_allow_html=True)

    latest_data = (
        df[[latest_q]].dropna().reset_index()
        .rename(columns={latest_q: "QoQ %", "Index": "Index"})
        .sort_values("QoQ %", ascending=False)
    )

    best_col, worst_col = st.columns(2)
    with best_col:
        st.markdown(f"<div style='font-size:11px;font-family:IBM Plex Mono,monospace;"
                    f"color:{RH_GREEN};margin-bottom:8px;'><b>↑ Top 5 this quarter</b></div>",
                    unsafe_allow_html=True)
        html = "<table class='val-table'><thead><tr><th>Index</th><th>QoQ %</th></tr></thead><tbody>"
        for _, r in latest_data.head(5).iterrows():
            html += f"<tr><td>{r['Index']}</td><td><span class='green'>{r['QoQ %']:+.1f}%</span></td></tr>"
        html += "</tbody></table>"
        st.markdown(html, unsafe_allow_html=True)

    with worst_col:
        st.markdown(f"<div style='font-size:11px;font-family:IBM Plex Mono,monospace;"
                    f"color:{RH_RED};margin-bottom:8px;'><b>↓ Bottom 5 this quarter</b></div>",
                    unsafe_allow_html=True)
        html = "<table class='val-table'><thead><tr><th>Index</th><th>QoQ %</th></tr></thead><tbody>"
        for _, r in latest_data.tail(5).iloc[::-1].iterrows():
            html += f"<tr><td>{r['Index']}</td><td><span class='red'>{r['QoQ %']:+.1f}%</span></td></tr>"
        html += "</tbody></table>"
        st.markdown(html, unsafe_allow_html=True)

    # ── Full 8-quarter table ──
    st.markdown(f"""<div class='section-title'>
        <span class='accent'>Table</span> &nbsp; Full 8-Quarter Grid ({metric_label})
    </div>""", unsafe_allow_html=True)

    html = "<table class='val-table'><thead><tr><th>Index</th>"
    for q in all_q_sorted:
        html += f"<th>{q}</th>"
    html += "</tr></thead><tbody>"
    for idx_name, row in df.iterrows():
        html += f"<tr><td>{idx_name}</td>"
        for q in all_q_sorted:
            v = row.get(q)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                html += "<td><span class='muted'>—</span></td>"
            elif abs(v) > CLAMP:
                html += f"<td><span style='color:#7B2FBE;font-weight:600;'>⚡{v:+.0f}%</span></td>"
            else:
                cls = "green" if v > 0 else "red"
                html += f"<td><span class='{cls}'>{v:+.1f}%</span></td>"
        html += "</tr>"
    html += "</tbody></table>"
    st.markdown(html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# VIEW 5 — Sector Rotation (quarterly heatmap, last 4 quarters)

def render_sector_rotation():
    st.markdown(f"""<div class='section-title'>
        <span class='accent'>Rotation</span> &nbsp; Sector Re-rating Momentum (Quarter-over-Quarter)
    </div>""", unsafe_allow_html=True)

    st.markdown(f"""<div style='font-size:11px;color:{RH_TEXT};background:{RH_SURFACE};
    border-left:3px solid {RH_GOLD};padding:12px 16px;margin-bottom:18px;line-height:1.6;'>
    <b style='color:{RH_MAROON};'>How to read this:</b> Unlike the main heatmap which shows
    cumulative re-rating, this shows the <b>change in P/E ratio quarter-over-quarter</b> —
    a momentum view. Where is money rotating <i>right now</i>?
    <br><br>
    &nbsp;&nbsp;<b style='color:{RH_RED};'>Red</b> = P/E expanded this quarter (sector got richer)
    <br>
    &nbsp;&nbsp;<b style='color:{RH_GREEN};'>Green</b> = P/E compressed this quarter (sector got cheaper)
    <br><br>
    Use this to spot rotation early — sectors flipping from cold to hot, or vice versa.
    </div>""", unsafe_allow_html=True)

    rows = []
    for ix in INDICES:
        snaps = data["indices"].get(ix["sheet_name"], {}).get("snapshots", [])
        if not snaps:
            continue
        # Get quarterly snapshots
        quarterly = quarterly_snapshots(snaps, start_year=2022)
        if len(quarterly) < 2:
            continue
        # Compute QoQ P/E change for each quarter (last 4 quarters max)
        row_data = {"Index": ix["display"]}
        for i in range(1, len(quarterly)):
            prev, curr = quarterly[i-1], quarterly[i]
            if prev.get("pe") and curr.get("pe") and prev["pe"] > 0:
                pe_chg = (curr["pe"] - prev["pe"]) / prev["pe"] * 100
                # Label as YYYY-Qn
                dt = pd.Timestamp(curr["date"])
                q = (dt.month - 1) // 3 + 1
                label = f"{dt.year}-Q{q}"
                row_data[label] = round(pe_chg, 1)
        if len(row_data) > 1:
            rows.append(row_data)

    if not rows:
        st.warning("Not enough quarterly P/E data yet. Run reconstruction script to backfill.")
        return

    df = pd.DataFrame(rows).set_index("Index")
    # Keep only the last 6 quarter columns (chronological)
    df = df[sorted(df.columns)]
    if len(df.columns) > 6:
        df = df.iloc[:, -6:]

    z = df.values
    text = [[f"{v:+.1f}%" if pd.notna(v) else "" for v in row] for row in z]

    fig = go.Figure(data=go.Heatmap(
        z=z, x=list(df.columns), y=list(df.index),
        text=text, texttemplate="%{text}",
        textfont={"family": "IBM Plex Mono", "size": 9, "color": "white"},
        colorscale=[
            [0.0,  RH_GREEN],
            [0.45, "#7FB97F"],
            [0.5,  "#FFFFFF"],
            [0.55, "#E78878"],
            [1.0,  RH_RED],
        ],
        zmid=0,
        hovertemplate="<b>%{y}</b><br>Quarter: %{x}<br>P/E change: %{z:+.1f}%<extra></extra>",
        colorbar=dict(title="P/E Δ%", tickfont=dict(color="#2C1810", family="IBM Plex Mono", size=10)),
    ))
    fig.update_layout(
        plot_bgcolor=RH_BG, paper_bgcolor=RH_SURFACE2,
        font=dict(family="IBM Plex Mono", color="#2C1810", size=10),
        height=max(500, 22 * len(rows)),
        margin=dict(l=140, r=40, t=20, b=40),
        xaxis=dict(tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810")),
        yaxis=dict(tickfont=dict(family="IBM Plex Mono", size=11, color="#2C1810")),
    )
    st.plotly_chart(fig, use_container_width=True)

    # ── Latest-quarter rotation summary ──
    if len(df.columns) >= 1:
        latest_col = df.columns[-1]
        st.markdown(f"""<div class='section-title'>
            <span class='accent'>Snapshot</span> &nbsp; Latest Quarter ({latest_col}) — Top Movers
        </div>""", unsafe_allow_html=True)

        latest = df[[latest_col]].dropna().rename(columns={latest_col: "PE Δ%"})
        latest = latest.reset_index().sort_values("PE Δ%", ascending=False)

        gain_col, loss_col = st.columns(2)
        with gain_col:
            st.markdown(f"<div style='font-size:11px;color:{RH_RED};margin-bottom:6px;'>"
                        f"<b>Got more expensive (P/E expanded)</b></div>", unsafe_allow_html=True)
            html = "<table class='val-table'><thead><tr><th>Index</th><th>PE Δ%</th></tr></thead><tbody>"
            for _, r in latest.head(5).iterrows():
                html += f"<tr><td>{r['Index']}</td><td><span class='red'>{r['PE Δ%']:+.1f}%</span></td></tr>"
            html += "</tbody></table>"
            st.markdown(html, unsafe_allow_html=True)
        with loss_col:
            st.markdown(f"<div style='font-size:11px;color:{RH_GREEN};margin-bottom:6px;'>"
                        f"<b>Got cheaper (P/E compressed)</b></div>", unsafe_allow_html=True)
            html = "<table class='val-table'><thead><tr><th>Index</th><th>PE Δ%</th></tr></thead><tbody>"
            for _, r in latest.tail(5).iloc[::-1].iterrows():
                html += f"<tr><td>{r['Index']}</td><td><span class='green'>{r['PE Δ%']:+.1f}%</span></td></tr>"
            html += "</tbody></table>"
            st.markdown(html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# VIEW 5 — Screener (pre-built filter buckets)
# ─────────────────────────────────────────────────────────────────────────────

def render_screener():
    st.markdown(f"""<div class='section-title'>
        <span class='accent'>Screener</span> &nbsp; Pre-Built Buckets
    </div>""", unsafe_allow_html=True)

    st.markdown(f"""<div style='font-size:11px;color:{RH_TEXT};background:{RH_SURFACE};
    border-left:3px solid {RH_GOLD};padding:12px 16px;margin-bottom:18px;line-height:1.6;'>
    Quick filters for a daily scan. Each bucket ranks all 25 indices by a single criterion.
    Click an index in any table to dig into its full view.
    </div>""", unsafe_allow_html=True)

    bucket = st.radio(
        "BUCKET",
        ["Cheapest right now", "Best earnings momentum", "Worst divergence", "Best de-rating opportunity"],
        horizontal=True, label_visibility="collapsed", key="screener_bucket",
    )

    # ── Build a master DataFrame with all the metrics ──
    rows = []
    for ix in INDICES:
        snaps = data["indices"].get(ix["sheet_name"], {}).get("snapshots", [])
        if not snaps:
            continue
        snaps_sorted = sorted(snaps, key=lambda s: s["date"])
        latest = snaps_sorted[-1]

        # PE z-score
        pe_snaps = [s["pe"] for s in snaps_sorted if s.get("pe") and s["pe"] > 0]
        pe_z = None
        pe_mean = None
        if len(pe_snaps) >= 4 and latest.get("pe"):
            pe_series = pd.Series(pe_snaps)
            pe_mean = pe_series.mean()
            pe_std = pe_series.std()
            if pe_std > 0:
                pe_z = (latest["pe"] - pe_mean) / pe_std

        # YoY EPS growth (latest vs ~4 quarters ago)
        latest_dt = pd.Timestamp(latest["date"])
        target_yago = latest_dt - pd.DateOffset(years=1)
        snap_dates = [pd.Timestamp(s["date"]) for s in snaps_sorted]
        deltas = [abs((d - target_yago).days) for d in snap_dates]
        yoy_eps = None
        if deltas and min(deltas) <= 180:
            base = snaps_sorted[deltas.index(min(deltas))]
            if base.get("eps") and latest.get("eps") and base["eps"] > 0:
                yoy_eps = (latest["eps"] - base["eps"]) / base["eps"] * 100

        # Spread since 2019 (cumulative CAGR price - eps)
        target_2019 = pd.Timestamp("2019-12-31")
        deltas_2019 = [abs((d - target_2019).days) for d in snap_dates]
        spread = None
        if deltas_2019 and min(deltas_2019) <= 180:
            base = snaps_sorted[deltas_2019.index(min(deltas_2019))]
            base_dt = pd.Timestamp(base["date"])
            yrs = (latest_dt - base_dt).days / 365.25
            if (yrs > 0.5 and base.get("close") and latest.get("close") and base["close"] > 0
                    and base.get("eps") and latest.get("eps") and base["eps"] > 0):
                px_cagr  = ((latest["close"] / base["close"]) ** (1/yrs) - 1) * 100
                eps_cagr = ((latest["eps"]   / base["eps"])   ** (1/yrs) - 1) * 100
                spread = px_cagr - eps_cagr

        rows.append({
            "Index":    ix["display"],
            "Sheet":    ix["sheet_name"],
            "PE":       latest.get("pe"),
            "PE Mean":  pe_mean,
            "PE Z":     pe_z,
            "YoY EPS":  yoy_eps,
            "Spread":   spread,
        })

    if not rows:
        st.warning("No data.")
        return

    df = pd.DataFrame(rows)

    # ── Render bucket ──
    if bucket == "Cheapest right now":
        sub = df.dropna(subset=["PE Z"]).sort_values("PE Z").head(10)
        st.markdown(f"<div style='font-size:11px;color:{RH_MUTED};margin-bottom:10px;'>"
                    f"Ranked by P/E z-score (current vs own historical mean). Negative = trading below average.</div>",
                    unsafe_allow_html=True)
        html = "<table class='val-table'><thead><tr><th>Rank</th><th>Index</th>" \
               "<th>Current P/E</th><th>Historical Mean</th><th>Z-score</th><th>Read</th></tr></thead><tbody>"
        for rank, (_, r) in enumerate(sub.iterrows(), 1):
            z = r["PE Z"]
            if z < -1: read, cls = "Cheap vs history", "green"
            elif z > 1: read, cls = "Expensive", "red"
            else:       read, cls = "Fair", "muted"
            html += f"<tr><td>{rank}</td><td>{r['Index']}</td>" \
                    f"<td>{r['PE']:.2f}</td><td>{r['PE Mean']:.2f}</td>" \
                    f"<td><span class='{cls}'>{z:+.2f}</span></td><td>{read}</td></tr>"
        html += "</tbody></table>"
        st.markdown(html, unsafe_allow_html=True)

    elif bucket == "Best earnings momentum":
        sub = df.dropna(subset=["YoY EPS"]).sort_values("YoY EPS", ascending=False).head(10)
        st.markdown(f"<div style='font-size:11px;color:{RH_MUTED};margin-bottom:10px;'>"
                    f"Ranked by year-over-year EPS growth. The earnings story sectors.</div>",
                    unsafe_allow_html=True)
        html = "<table class='val-table'><thead><tr><th>Rank</th><th>Index</th>" \
               "<th>YoY EPS Growth</th><th>Current P/E</th></tr></thead><tbody>"
        for rank, (_, r) in enumerate(sub.iterrows(), 1):
            y = r["YoY EPS"]
            cls = "green" if y > 0 else "red"
            pe_str = f"{r['PE']:.2f}" if pd.notna(r["PE"]) else "—"
            html += f"<tr><td>{rank}</td><td>{r['Index']}</td>" \
                    f"<td><span class='{cls}'>{y:+.1f}%</span></td><td>{pe_str}</td></tr>"
        html += "</tbody></table>"
        st.markdown(html, unsafe_allow_html=True)

    elif bucket == "Worst divergence":
        sub = df.dropna(subset=["Spread"]).sort_values("Spread", ascending=False).head(10)
        st.markdown(f"<div style='font-size:11px;color:{RH_MUTED};margin-bottom:10px;'>"
                    f"Indices where price compounded most ahead of earnings since 2019. "
                    f"Potential mean-reversion candidates.</div>",
                    unsafe_allow_html=True)
        html = "<table class='val-table'><thead><tr><th>Rank</th><th>Index</th>" \
               "<th>Spread (since 2019)</th><th>Current P/E</th></tr></thead><tbody>"
        for rank, (_, r) in enumerate(sub.iterrows(), 1):
            s = r["Spread"]
            pe_str = f"{r['PE']:.2f}" if pd.notna(r["PE"]) else "—"
            html += f"<tr><td>{rank}</td><td>{r['Index']}</td>" \
                    f"<td><span class='red'>{s:+.1f}%</span></td><td>{pe_str}</td></tr>"
        html += "</tbody></table>"
        st.markdown(html, unsafe_allow_html=True)

    elif bucket == "Best de-rating opportunity":
        sub = df.dropna(subset=["Spread"]).sort_values("Spread", ascending=True).head(10)
        st.markdown(f"<div style='font-size:11px;color:{RH_MUTED};margin-bottom:10px;'>"
                    f"Indices where earnings compounded most ahead of price since 2019. "
                    f"P/E has compressed — potentially undervalued.</div>",
                    unsafe_allow_html=True)
        html = "<table class='val-table'><thead><tr><th>Rank</th><th>Index</th>" \
               "<th>Spread (since 2019)</th><th>Current P/E</th></tr></thead><tbody>"
        for rank, (_, r) in enumerate(sub.iterrows(), 1):
            s = r["Spread"]
            pe_str = f"{r['PE']:.2f}" if pd.notna(r["PE"]) else "—"
            html += f"<tr><td>{rank}</td><td>{r['Index']}</td>" \
                    f"<td><span class='green'>{s:+.1f}%</span></td><td>{pe_str}</td></tr>"
        html += "</tbody></table>"
        st.markdown(html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

def build_excel():
    """Rebuild the original .xlsx structure from JSON. One sheet per index."""
    wb = Workbook()
    wb.remove(wb.active)

    maroon_fill = PatternFill('solid', start_color='FF8B1A1A')
    gold_fill   = PatternFill('solid', start_color='FFD4A830')
    cream_fill  = PatternFill('solid', start_color='FFEDE2C8')
    white_font  = Font(name='Arial', color='FFFFFFFF', bold=True, size=10)
    dark_font   = Font(name='Arial', color='FF2C1810', size=10)
    bold_font   = Font(name='Arial', color='FF8B1A1A', bold=True, size=11)

    thin = Side(border_style='thin', color='FFD4A830')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for ix in INDICES:
        sheet = wb.create_sheet(ix["sheet_name"][:31])
        snaps = data["indices"].get(ix["sheet_name"], {}).get("snapshots", [])
        if not snaps:
            sheet['B3'] = f"{ix['display']} — no data"
            continue

        years = list(range(ix.get("start_year", 2012), datetime.now().year))
        yearly = yearly_snapshots(snaps, years)
        yr_rows = compute_yearly_growth(yearly)
        quarterly = quarterly_snapshots(snaps, start_year=ix.get("start_year", 2012))
        q_rows = compute_quarterly_growth(quarterly)

        # ── Yearly block ──
        sheet['B3'] = f"{ix['display']} Yearly Performance"
        sheet['B3'].font = bold_font

        # Header row 4: dates
        for i, r in enumerate(yr_rows):
            c = sheet.cell(row=4, column=3+i, value=r["date"])
            c.fill = maroon_fill; c.font = white_font
            c.alignment = Alignment(horizontal='center')

        # Rows 5-9: close, PB, EPS, PE, RSI
        metric_rows = [
            (5, ix["display"], "close"),
            (6, "PB",          "pb"),
            (7, "EPS",         "eps"),
            (8, "PE",          "pe"),
            (9, "RSI",         "rsi"),
        ]
        for row_n, label, key in metric_rows:
            cl = sheet.cell(row=row_n, column=2, value=label)
            cl.fill = cream_fill; cl.font = Font(name='Arial', color='FFA8741A', bold=True, size=10)
            for i, r in enumerate(yr_rows):
                v = r.get(key)
                sheet.cell(row=row_n, column=3+i, value=v if v is not None else "")

        # Row 11: years
        sheet.cell(row=11, column=2, value="No. of Years").font = dark_font
        for i, r in enumerate(yr_rows[1:], start=1):
            sheet.cell(row=11, column=3+i, value=r["years"])

        # Rows 12-15: CAGR
        cagr_rows = [
            (12, f"{ix['display']} YOY CAGR", "yoy_idx"),
            (13, "EPS YOY CAGR",              "yoy_eps"),
            (14, f"{ix['display']} CUM CAGR", "cum_idx"),
            (15, "EPS CUM CAGR",              "cum_eps"),
        ]
        for row_n, label, key in cagr_rows:
            cl = sheet.cell(row=row_n, column=2, value=label)
            cl.fill = cream_fill; cl.font = Font(name='Arial', color='FFA8741A', bold=True, size=10)
            for i, r in enumerate(yr_rows[1:], start=1):
                v = r.get(key)
                cell = sheet.cell(row=row_n, column=3+i, value=v if v is not None else "")
                cell.number_format = '0.00%'

        # ── Quarterly block ──
        sheet['B18'] = f"{ix['display']} Quarterly Performance"
        sheet['B18'].font = bold_font

        for i, r in enumerate(q_rows):
            c = sheet.cell(row=19, column=3+i, value=r["date"])
            c.fill = maroon_fill; c.font = white_font
            c.alignment = Alignment(horizontal='center')

        q_metric = [(20, ix["display"], "close"), (21, "EPS", "eps")]
        for row_n, label, key in q_metric:
            cl = sheet.cell(row=row_n, column=2, value=label)
            cl.fill = cream_fill; cl.font = Font(name='Arial', color='FFA8741A', bold=True, size=10)
            for i, r in enumerate(q_rows):
                v = r.get(key)
                sheet.cell(row=row_n, column=3+i, value=v if v is not None else "")

        q_growth = [
            (22, "QOQ Gr (Index)", "qoq_idx"),
            (23, "QOQ Gr (EPS)",   "qoq_eps"),
            (24, "YOY Gr (Index)", "yoy_idx"),
            (25, "YOY Gr (EPS)",   "yoy_eps"),
        ]
        for row_n, label, key in q_growth:
            cl = sheet.cell(row=row_n, column=2, value=label)
            cl.fill = cream_fill; cl.font = Font(name='Arial', color='FFA8741A', bold=True, size=10)
            for i, r in enumerate(q_rows):
                v = r.get(key)
                cell = sheet.cell(row=row_n, column=3+i, value=v if v is not None else "")
                cell.number_format = '0.00'

        # Column widths
        sheet.column_dimensions['B'].width = 28
        for col in range(3, 3 + max(len(yr_rows), len(q_rows)) + 2):
            sheet.column_dimensions[chr(64+col)].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# Router
# ─────────────────────────────────────────────────────────────────────────────

if view == "Single Index":
    render_single_index(selected_sheet)
elif view == "Cross-Index Heatmap":
    render_heatmap()
elif view == "Divergence Chart":
    render_divergence()
elif view == "QoQ Growth":
    render_qoq_heatmap()
elif view == "Sector Rotation":
    render_sector_rotation()
elif view == "Screener":
    render_screener()

st.markdown(f"<div style='margin-top:30px;'></div>", unsafe_allow_html=True)

exp_col, _ = st.columns([1, 4])
with exp_col:
    if st.button("📥 EXPORT TO EXCEL", use_container_width=True):
        buf = build_excel()
        st.download_button(
            "DOWNLOAD .XLSX",
            data=buf,
            file_name=f"Price_Vs_EPS_Growth_Comparison_{datetime.now():%b%y}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

st.markdown(f"""<div style='font-size:9px;color:{RH_MUTED};margin-top:30px;text-align:center;'>
Data sources: NSE daily index archives (P/E, P/B), yfinance (prices, RSI). EPS derived as Close ÷ P/E.
Daily auto-update Mon-Fri 4PM IST via GitHub Actions.
</div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)
if st.button("← BACK TO HUB", use_container_width=True):
    st.switch_page("Home.py")
